import streamlit as st
import pandas as pd
import os
import re
import math
import base64
import requests
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from streamlit_echarts import st_echarts, JsCode 

# ---------------------------------------------------------
# ログイン（パスワード認証）
# ---------------------------------------------------------
PASSWORD = st.secrets["PASSWORD"]

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    pw = st.text_input("パスワードを入力してください", type="password")
    if pw == PASSWORD:
        st.session_state.authenticated = True
        st.rerun()
    elif pw != "0128":
        st.error("パスワードが違います")
    st.stop()

# ---------------------------------------------------------
# GitHub secrets 読み込み
# ---------------------------------------------------------
GITHUB_TOKEN = st.secrets["GITHUB_TOKEN"]
GITHUB_REPO = st.secrets["GITHUB_REPO"]
GITHUB_FILE_PATH = st.secrets["GITHUB_FILE_PATH"]

# ---------------------------------------------------------
# GitHub から Excel を取得
# ---------------------------------------------------------
def download_excel_from_github(repo, file_path, token, local_path="temp.xlsx"):
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"
    headers = {"Authorization": f"token {token}"}

    res = requests.get(url, headers=headers)
    if res.status_code == 200:
        content = base64.b64decode(res.json()["content"])
        with open(local_path, "wb") as f:
            f.write(content)
        return local_path
    else:
        st.error("GitHub からファイルを取得できませんでした")
        return None

# ---------------------------------------------------------
# GitHub へ Excel をアップロード
# ---------------------------------------------------------
def update_excel_to_github(local_path, repo, file_path, token, commit_message="Update Excel"):
    url = f"https://api.github.com/repos/{repo}/contents/{file_path}"

    with open(local_path, "rb") as f:
        content = f.read()
    encoded = base64.b64encode(content).decode()

    res = requests.get(url, headers={"Authorization": f"token {token}"})
    sha = res.json().get("sha", None)

    data = {
        "message": commit_message,
        "content": encoded,
        "sha": sha
    }

    res = requests.put(url, json=data, headers={"Authorization": f"token {token}"})
    return res.status_code in [200, 201]

# ---------------------------------------------------------
# 他のシートを消さずに、指定シートだけ更新する関数
# ---------------------------------------------------------
def save_sheet_without_deleting_others(excel_path, sheet_name, df):
    wb = load_workbook(excel_path)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)

    ws = wb.create_sheet(sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(excel_path)

# ---------------------------------------------------------
# 列名を正規化
# ---------------------------------------------------------
def normalize_columns(df):
    new_cols = []
    for col in df.columns:
        c = str(col)
        c = c.replace(" ", "").replace("　", "")
        c = c.replace("ヒヅケ", "日付")
        new_cols.append(c)
    df.columns = new_cols
    return df

# ---------------------------------------------------------
# 競泳表記 → 秒
# ---------------------------------------------------------
def time_to_seconds(t):
    if t is None:
        return None

    if isinstance(t, pd.Timestamp):
        return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1e6

    if isinstance(t, (int, float)) and t > 30000:
        return None

    if isinstance(t, (int, float)):
        if 0 < t < 1:
            return t * 86400
        else:
            return float(t)

    s = str(t).strip()
    s = s.replace("：", ":")

    m = re.match(r"(\d+)'(\d+)[\"”]?(\d+)", s)
    if m:
        minutes = int(m.group(1))
        seconds = int(m.group(2))
        ms = int(m.group(3))
        return minutes * 60 + seconds + ms / 100

    if ":" in s:
        try:
            m, sec = s.split(":")
            return int(m) * 60 + float(sec)
        except:
            pass

    try:
        return float(s)
    except:
        return None

# ---------------------------------------------------------
# 秒 → 競泳表記
# ---------------------------------------------------------
def seconds_to_swim_format(sec):
    if sec is None or (isinstance(sec, float) and math.isnan(sec)):
        return "―"
    m = int(sec // 60)
    s = sec % 60
    return f"{m}'{s:05.2f}"

# ---------------------------------------------------------
# GitHub から最新 Excel を取得
# ---------------------------------------------------------
local_excel = download_excel_from_github(GITHUB_REPO, GITHUB_FILE_PATH, GITHUB_TOKEN)

if local_excel is None:
    st.stop()

# ---------------------------------------------------------
# 種目選択
# ---------------------------------------------------------
events = ["フリー", "バッタ", "ブレ", "バック", "メドレー"]
event = st.selectbox("種目を選択してください", events)

# ---------------------------------------------------------
# 固定ヘッダー
# ---------------------------------------------------------
event_colors = {
    "フリー": "#1E90FF",
    "バッタ": "#FF8C00",
    "ブレ":   "#32CD32",
    "バック": "#8A2BE2",
    "メドレー": "#DC143C"
}

header_color = event_colors.get(event, "#000000")

st.markdown(
    f"""
    <style>
        .header-title {{
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            background-color: {header_color};
            padding: 18px 20px;
            font-size: 32px;
            font-weight: bold;
            color: white;
            text-align: center;
            border-bottom: 3px solid #ddd;
            z-index: 999999;
            display: flex;
            justify-content: center;
            gap: 40px;
        }}

        @media screen and (max-width: 600px) {{
            .header-title {{
                font-size: 20px;
                padding: 12px 10px;
                gap: 20px;
            }}
            .block-container {{
                padding-top: 100px !important;
            }}
        }}

        .block-container {{
            padding-top: 140px !important;
        }}
    </style>

    <div class="header-title">
        <span>HONOKA Swimming record</span>
        <span>{event}</span>
    </div>
    """,
    unsafe_allow_html=True
)

# ---------------------------------------------------------
# Excel 読み込み
# ---------------------------------------------------------
sheet_name = event

data = pd.read_excel(local_excel, sheet_name=sheet_name)
data = data.iloc[:, :6]
data.columns = ["日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]
data = normalize_columns(data)

data["タイム"] = data["タイム"].apply(time_to_seconds)

data["距離"] = pd.to_numeric(data["距離"], errors="coerce")
data = data.dropna(subset=["距離"])
data["距離"] = data["距離"].astype(int)

# ---------------------------------------------------------
# 距離選択
# ---------------------------------------------------------
if event == "メドレー":
    distance_list = [200, 400]
elif event == "ブレ":
    distance_list = [50, 100]
else:
    distance_list = sorted(data["距離"].unique())

distance = st.selectbox("距離を選択してください", distance_list)

# ---------------------------------------------------------
# 長水路／短水路
# ---------------------------------------------------------
course = st.selectbox("長水路／短水路を選択", ["全記録", "短水路", "長水路"])

# ---------------------------------------------------------
# データ絞り込み
# ---------------------------------------------------------
if course == "全記録":
    filtered = data[data["距離"] == distance].sort_values("日付")
else:
    filtered = data[
        (data["距離"] == distance) &
        (data["長水路or短水路"] == course)
    ].sort_values("日付")

filtered = filtered[filtered["タイム"].notna()]

if filtered.empty:
    st.error(f"{event} の {distance}m（{course}）のデータがありません")
    st.stop()

# ---------------------------------------------------------
# ECharts 用データ準備
# ---------------------------------------------------------
filtered["日付_学年"] = (
    filtered["日付"].dt.strftime("%Y-%m-%d") + "（" + filtered["学年"] + "）"
)

filtered["タイム_表示"] = filtered["タイム"].apply(seconds_to_swim_format)

x_data = filtered["日付_学年"].tolist()
y_data = filtered["タイム"].tolist()
y_label = filtered["タイム_表示"].tolist()

# ---------------------------------------------------------
# Y軸レンジ（メドレーは10秒刻み、他は2秒刻み）
# ---------------------------------------------------------
y_min_raw = min(y_data)
y_max_raw = max(y_data)

if "メドレー" in event:
    y_min = math.floor(y_min_raw / 10) * 10
    y_max = math.ceil(y_max_raw / 10) * 10
    y_interval = 10
else:
    y_min = math.floor(y_min_raw / 2) * 2
    y_max = math.ceil(y_max_raw / 2) * 2
    y_interval = 2

from streamlit_echarts import st_echarts, JsCode

# ---------------------------------------------------------
# 点の色分け（長水路＝青、短水路＝赤）
# ---------------------------------------------------------
series_data = [
    {
        "value": y_data[i],
        "label": y_label[i],
        "itemStyle": {
            "color": "#3366FF" if filtered["長水路or短水路"].iloc[i] == "長水路" else "#FF3333"
        }
    }
    for i in range(len(y_data))
]

# ---------------------------------------------------------
# Y軸フォーマッタ（メドレーは分＋秒）
# ---------------------------------------------------------
if "メドレー" in event:
    y_axis_formatter = JsCode("""
        function (value) {
            var min = Math.floor(value / 60);
            var sec = value % 60;
            return min + "'" + sec.toFixed(2).padStart(5, '0');
        }
    """)
else:
    y_axis_formatter = "{value}"

# ---------------------------------------------------------
# ECharts オプション
# ---------------------------------------------------------
options = {
    "title": {
        "text": f"{event} {distance}m（{course}）の記録推移"
    },

    # ★ 凡例をタイトルの上に配置
    "legend": {
        "top": 0,
        "left": "center",
        "data": ["長水路", "短水路"],
        "textStyle": {"color": "#000"}
    },

    "tooltip": {
        "trigger": "axis",
        "formatter": JsCode("""
            function (params) {
                return params[0].data.label;
            }
        """)
    },

    "xAxis": {
        "type": "category",
        "data": x_data
    },

    "yAxis": {
        "type": "value",
        "inverse": False,
        "min": y_min,
        "max": y_max,
        "interval": y_interval,
        "axisLabel": {
            "formatter": y_axis_formatter
        }
    },

    "dataZoom": [
        {"type": "inside"},
        {"type": "slider"}
    ],

    "series": [
        # ★ ダミー凡例（青）
        {
            "name": "長水路",
            "type": "line",
            "data": [],
            "lineStyle": {"color": "#3366FF"},
            "showSymbol": True,
            "symbol": "circle",
            "symbolSize": 12
        },
        # ★ ダミー凡例（赤）
        {
            "name": "短水路",
            "type": "line",
            "data": [],
            "lineStyle": {"color": "#FF3333"},
            "showSymbol": True,
            "symbol": "circle",
            "symbolSize": 12
        },
        # ★ 実データ（線は灰色、点は青/赤）
        {
            "type": "line",
            "data": series_data,
            "smooth": False,
            "lineStyle": {"color": "gray", "width": 2},
            "label": {
                "show": True,
                "position": "top",
                "formatter": JsCode("function (p) { return p.data.label; }"),
                "fontSize": 12
            }
        }
    ]
}

# ---------------------------------------------------------
# グラフ描画
# ---------------------------------------------------------
st_echarts(options=options, height="500px")

# ---------------------------------------------------------
# ベストタイム
# ---------------------------------------------------------
best_short = data[(data["距離"] == distance) & (data["長水路or短水路"] == "短水路") & (data["タイム"].notna())]
best_long  = data[(data["距離"] == distance) & (data["長水路or短水路"] == "長水路") & (data["タイム"].notna())]

st.subheader("ベストタイム（短水路）")
if not best_short.empty:
    t = best_short["タイム"].min()
    d = best_short.loc[best_short["タイム"].idxmin(), "日付"]
    st.write(f"ベストタイム：**{seconds_to_swim_format(t)}**")
    st.write(f"更新日：{d}")
else:
    st.write("データなし")

st.subheader("ベストタイム（長水路）")
if not best_long.empty:
    t = best_long["タイム"].min()
    d = best_long.loc[best_long["タイム"].idxmin(), "日付"]
    st.write(f"ベストタイム：**{seconds_to_swim_format(t)}**")
    st.write(f"更新日：{d}")
else:
    st.write("データなし")

# ---------------------------------------------------------
# 新しい記録を追加
# ---------------------------------------------------------
st.subheader("新しい記録を追加")

with st.form("add_record_form"):
    new_date = st.date_input("日付")
    new_grade = st.selectbox("学年", ["小6","中1","中2","中3"])
    new_distance = st.selectbox("距離", distance_list)
    new_course = st.selectbox("長水路 or 短水路", ["長水路", "短水路"])
    new_time_str = st.text_input(
        "タイム（入力方法）\n\n"
        "【60秒未満】例：58秒11 → 58.11\n"
        "【60秒以上】例：1分41秒58 → 1'41\"58\n\n"
        "※ どちらの形式でも自動で変換されます"
    )
    new_place = st.text_input("会場", value="菰野スイミング")

    submitted = st.form_submit_button("追加する")

if submitted:
    new_time_sec = time_to_seconds(new_time_str)

    if new_time_sec is None:
        st.error("タイムの形式が正しくありません")
    else:
        new_row = pd.DataFrame([{
            "日付": pd.to_datetime(new_date),
            "学年": new_grade,
            "距離": int(new_distance),
            "長水路or短水路": new_course,
            "タイム": new_time_sec,
            "会場": new_place
        }])

        try:
            book = pd.read_excel(local_excel, sheet_name=sheet_name)
            book = normalize_columns(book)
            book = book.iloc[:, :6]
            book.columns = ["日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]

            updated = pd.concat([book, new_row], ignore_index=True)

            save_sheet_without_deleting_others(local_excel, sheet_name, updated)

            update_excel_to_github(
                local_path=local_excel,
                repo=GITHUB_REPO,
                file_path=GITHUB_FILE_PATH,
                token=GITHUB_TOKEN,
                commit_message=f"Add record: {event} {distance}m"
            )

            st.success("記録を追加しました！（GitHub にも反映済み）")
            st.rerun()

        except Exception as e:
            st.error(f"Excel 書き込みエラー: {e}")

# ---------------------------------------------------------
# 記録の修正・削除
# ---------------------------------------------------------
st.subheader("記録の修正・削除")

edit_df = filtered.copy().reset_index(drop=True)
edit_df["行番号"] = edit_df.index

st.dataframe(edit_df[["行番号", "日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]])

target_index = st.number_input("修正・削除する行番号を入力", min_value=0, max_value=len(edit_df)-1, step=1)

target_row = edit_df.iloc[target_index]

st.write("選択中の記録：")
st.write(target_row)

# -------------------------
# 修正フォーム
# -------------------------
with st.form("edit_form"):
    e_date = st.date_input("日付（修正）", value=target_row["日付"])
    e_grade = st.selectbox("学年（修正）", ["小1","小2","小3","小4","小5","小6","中1","中2","中3"],
                           index=["小1","小2","小3","小4","小5","小6","中1","中2","中3"].index(target_row["学年"]))
    e_distance = st.number_input("距離（修正）", value=int(target_row["距離"]))
    e_course = st.selectbox("長水路 or 短水路（修正）", ["長水路", "短水路"],
                            index=0 if target_row["長水路or短水路"]=="長水路" else 1)
    e_time_str = st.text_input("タイム（修正）", value=seconds_to_swim_format(target_row["タイム"]))
    e_place = st.text_input("会場（修正）", value=target_row["会場"])

    edit_submitted = st.form_submit_button("修正する")

# -------------------------
# 修正処理
# -------------------------
if edit_submitted:
    new_time_sec = time_to_seconds(e_time_str)

    if new_time_sec is None:
        st.error("タイムの形式が正しくありません")
    else:
        book = pd.read_excel(local_excel, sheet_name=sheet_name)
        book = normalize_columns(book)
        book = book.iloc[:, :6]
        book.columns = ["日付", "学年", "距離", "長水路or短水路", "タイム", "会場"]

        book.loc[target_row.name] = [
            pd.to_datetime(e_date),
            e_grade,
            int(e_distance),
            e_course,
            new_time_sec,
            e_place
        ]

        save_sheet_without_deleting_others(local_excel, sheet_name, book)

        update_excel_to_github(
            local_path=local_excel,
            repo=GITHUB_REPO,
            file_path=GITHUB_FILE_PATH,
            token=GITHUB_TOKEN,
            commit_message=f"Edit record: {event} {distance}m"
        )

        st.success("修正しました！（GitHub にも反映済み）")
        st.rerun()